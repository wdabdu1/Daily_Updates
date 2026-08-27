import io
from datetime import date, datetime
from decimal import InvalidOperation

import openpyxl
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from .. import models, schemas
from ..auth import get_current_user, require_write
from ..config import NON_SDG_RATE_TYPES, SDG_RATE_TYPES
from ..database import get_db
from ..export_utils import parse_decimal, read_uploaded_xlsx, xlsx_response, xlsx_template_response
from ..rates import build_daily_series, get_pair

router = APIRouter(prefix="/api/fx", tags=["fx"])

FX_TEMPLATE_COLUMNS = ["Date", "Base Currency", "Quote Currency", "Rate Type", "Rate"]

# The three currencies that always get Market/CBOS/Pricing rates against
# SDG -- the "compulsory 3-currency batch" the FX Rates page is built
# around. Order matches how they're entered on the page (USD, Euro, AED).
CORE_FX_CURRENCIES = ["USD", "EUR", "AED"]

# str.title() mangles acronyms ("CBOS".title() == "Cbos"), so rate types
# from an upload are normalized via this explicit case-insensitive map
# instead, rather than relying on .title()/.capitalize().
_RATE_TYPE_CANONICAL = {t.lower(): t for t in SDG_RATE_TYPES}


def _normalize_rate_type(raw: str) -> str | None:
    return _RATE_TYPE_CANONICAL.get(str(raw).strip().lower())


def _get_or_create_sdg_pair(db: Session, currency: str) -> models.CurrencyPair:
    """Look up (or silently create, same as the generic importer does) the
    CURRENCY/SDG pair. Every currency in CORE_FX_CURRENCIES is expected to
    be priced against SDG here -- this is the pair the compulsory
    Market/CBOS/Pricing batch entries are written to."""
    pair = get_pair(db, currency, "SDG")
    if pair:
        return pair
    pair = models.CurrencyPair(
        base_currency=currency, quote_currency="SDG", supports_extended_rates=True, is_default=False
    )
    db.add(pair)
    db.flush()
    return pair


def _upsert_rate(
    db: Session, pair: models.CurrencyPair, rate_type: str, rate_date: date, rate
) -> tuple[models.FxRate, bool]:
    """Create or update the (pair, rate_type, rate_date) row. Returns
    (row, was_update)."""
    existing = (
        db.query(models.FxRate)
        .filter(
            models.FxRate.rate_date == rate_date,
            models.FxRate.currency_pair_id == pair.id,
            models.FxRate.rate_type == rate_type,
        )
        .first()
    )
    if existing:
        existing.rate = rate
        existing.is_manual_entry = True
        return existing, True
    row = models.FxRate(
        rate_date=rate_date, currency_pair_id=pair.id, rate_type=rate_type, rate=rate, is_manual_entry=True
    )
    db.add(row)
    db.flush()
    return row, False


@router.get("/rate-types")
def rate_types(pair_id: int, db: Session = Depends(get_db), _u: models.User = Depends(get_current_user)):
    pair = db.query(models.CurrencyPair).get(pair_id)
    if not pair:
        raise HTTPException(404, "Unknown currency pair.")
    return SDG_RATE_TYPES if pair.supports_extended_rates else NON_SDG_RATE_TYPES


@router.post("/rates", response_model=schemas.FxRateOut)
def record_rate(
    payload: schemas.FxRateCreate,
    db: Session = Depends(get_db),
    _u: models.User = Depends(require_write),
):
    pair = db.query(models.CurrencyPair).get(payload.currency_pair_id)
    if not pair:
        raise HTTPException(400, "Unknown currency pair.")
    allowed_types = SDG_RATE_TYPES if pair.supports_extended_rates else NON_SDG_RATE_TYPES
    if payload.rate_type not in allowed_types:
        raise HTTPException(
            400, f"{pair.label} only supports rate types: {', '.join(allowed_types)}"
        )

    existing = (
        db.query(models.FxRate)
        .filter(
            models.FxRate.rate_date == payload.rate_date,
            models.FxRate.currency_pair_id == payload.currency_pair_id,
            models.FxRate.rate_type == payload.rate_type,
        )
        .first()
    )
    if existing:
        existing.rate = payload.rate
        existing.is_manual_entry = True
    else:
        existing = models.FxRate(
            rate_date=payload.rate_date,
            currency_pair_id=payload.currency_pair_id,
            rate_type=payload.rate_type,
            rate=payload.rate,
            is_manual_entry=True,
        )
        db.add(existing)
    db.commit()
    db.refresh(existing)
    return schemas.FxRateOut(
        id=existing.id,
        rate_date=existing.rate_date,
        currency_pair=pair.label,
        rate_type=existing.rate_type,
        rate=existing.rate,
        is_carried_forward=False,
    )


@router.post("/rates/batch", response_model=schemas.FxBatchRateOut)
def record_rate_batch(
    payload: schemas.FxBatchRateCreate,
    db: Session = Depends(get_db),
    _u: models.User = Depends(require_write),
):
    """Save Market, CBOS, or Pricing for USD, Euro and AED (all vs SDG) in
    one all-or-nothing save -- this is the primary way rates get entered
    now. Pydantic already made all three values compulsory; this just
    writes the three FxRate rows together and rolls back together if
    anything goes wrong."""
    if payload.rate_type not in SDG_RATE_TYPES:
        raise HTTPException(400, f"Rate type must be one of: {', '.join(SDG_RATE_TYPES)}")

    try:
        usd_pair = _get_or_create_sdg_pair(db, "USD")
        eur_pair = _get_or_create_sdg_pair(db, "EUR")
        aed_pair = _get_or_create_sdg_pair(db, "AED")
        usd_row, _ = _upsert_rate(db, usd_pair, payload.rate_type, payload.rate_date, payload.usd_rate)
        eur_row, _ = _upsert_rate(db, eur_pair, payload.rate_type, payload.rate_date, payload.euro_rate)
        aed_row, _ = _upsert_rate(db, aed_pair, payload.rate_type, payload.rate_date, payload.aed_rate)
    except Exception:
        db.rollback()
        raise
    db.commit()
    db.refresh(usd_row)
    db.refresh(eur_row)
    db.refresh(aed_row)

    def _out(row: models.FxRate, pair: models.CurrencyPair) -> schemas.FxRateOut:
        return schemas.FxRateOut(
            id=row.id,
            rate_date=row.rate_date,
            currency_pair=pair.label,
            rate_type=row.rate_type,
            rate=row.rate,
            is_carried_forward=False,
        )

    return schemas.FxBatchRateOut(
        rate_date=payload.rate_date,
        rate_type=payload.rate_type,
        usd=_out(usd_row, usd_pair),
        euro=_out(eur_row, eur_pair),
        aed=_out(aed_row, aed_pair),
    )


@router.patch("/rates/{rate_id}", response_model=schemas.FxRateOut)
def update_rate(
    rate_id: int,
    payload: schemas.FxRateUpdate,
    db: Session = Depends(get_db),
    _u: models.User = Depends(require_write),
):
    rate_row = db.query(models.FxRate).get(rate_id)
    if not rate_row:
        raise HTTPException(404, "Rate entry not found.")
    rate_row.rate = payload.rate
    rate_row.is_manual_entry = True
    db.commit()
    db.refresh(rate_row)
    pair = db.query(models.CurrencyPair).get(rate_row.currency_pair_id)
    return schemas.FxRateOut(
        id=rate_row.id,
        rate_date=rate_row.rate_date,
        currency_pair=pair.label if pair else "",
        rate_type=rate_row.rate_type,
        rate=rate_row.rate,
        is_carried_forward=False,
    )


@router.delete("/rates/{rate_id}")
def delete_rate(
    rate_id: int, db: Session = Depends(get_db), _u: models.User = Depends(require_write)
):
    rate_row = db.query(models.FxRate).get(rate_id)
    if not rate_row:
        raise HTTPException(404, "Rate entry not found.")
    db.delete(rate_row)
    db.commit()
    return {"deleted": True}


@router.get("/rates/table", response_model=list[schemas.FxRateOut])
def rates_table(
    currency_pair_id: int,
    rate_type: str = "Market",
    start: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    _u: models.User = Depends(get_current_user),
):
    """Daily series with carry-forward fill, used both for the FX Rates
    page's current-month table (and collapsed prior months) and for
    Analysis charts."""
    pair = db.query(models.CurrencyPair).get(currency_pair_id)
    if not pair:
        raise HTTPException(404, "Unknown currency pair.")
    series = build_daily_series(db, pair.base_currency, pair.quote_currency, rate_type, start, end)
    return [
        schemas.FxRateOut(
            id=row["id"],
            rate_date=row["rate_date"],
            currency_pair=pair.label,
            rate_type=rate_type,
            rate=row["rate"],
            is_carried_forward=row["is_carried_forward"],
        )
        for row in series
    ]


def _combined_rows(db: Session, currency: str, start: date, end: date) -> list[schemas.FxCombinedRow]:
    """Market/CBOS/Pricing side by side per calendar day for one currency
    (vs SDG). Shared by the /rates/combined JSON endpoint (Rate History
    table) and the /rates/combined/export Excel download (Round 13) so the
    two always agree on exactly what a given currency/date-window contains."""
    currency = currency.strip().upper()
    pair = get_pair(db, currency, "SDG")
    if not pair:
        return []

    by_type: dict[str, dict[date, dict]] = {}
    for rt in SDG_RATE_TYPES:
        series = build_daily_series(db, currency, "SDG", rt, start, end)
        by_type[rt] = {row["rate_date"]: row for row in series}

    all_dates = sorted(set().union(*(s.keys() for s in by_type.values())))
    rows = []
    for d in all_dates:
        m = by_type["Market"].get(d)
        c = by_type["CBOS"].get(d)
        p = by_type["Pricing"].get(d)
        rows.append(
            schemas.FxCombinedRow(
                rate_date=d,
                market_rate=m["rate"] if m else None,
                market_id=m["id"] if m else None,
                market_carried_forward=m["is_carried_forward"] if m else False,
                cbos_rate=c["rate"] if c else None,
                cbos_id=c["id"] if c else None,
                cbos_carried_forward=c["is_carried_forward"] if c else False,
                pricing_rate=p["rate"] if p else None,
                pricing_id=p["id"] if p else None,
                pricing_carried_forward=p["is_carried_forward"] if p else False,
            )
        )
    return rows


@router.get("/rates/combined", response_model=list[schemas.FxCombinedRow])
def rates_combined(
    currency: str,
    start: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    _u: models.User = Depends(get_current_user),
):
    """Powers the redesigned FX Rates page: for one selected currency (vs
    SDG), Market/CBOS/Pricing side by side per calendar day. Averaging and
    period filtering happen client-side over this series so the same data
    can drive "All / Quarter / Month" without extra round-trips."""
    return _combined_rows(db, currency, start, end)


@router.get("/rates/combined/export")
def export_rates_combined(
    currency: str,
    start: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    _u: models.User = Depends(get_current_user),
):
    """Excel export of the Rate History table (Round 13) -- the frontend
    passes whatever currency/date-window its All/Quarter/Month filter
    currently resolves to, so the download always matches what's on
    screen rather than dumping the full unfiltered history."""
    currency = currency.strip().upper()
    rows = _combined_rows(db, currency, start, end)
    out_rows = [
        {
            "Date": r.rate_date,
            "Currency": currency,
            "Market": float(r.market_rate) if r.market_rate is not None else None,
            "CBOS": float(r.cbos_rate) if r.cbos_rate is not None else None,
            "Pricing": float(r.pricing_rate) if r.pricing_rate is not None else None,
        }
        for r in rows
    ]
    return xlsx_response(out_rows, f"fx_rate_history_{currency}.xlsx")


@router.get("/rates/table/export")
def export_rates_table(
    currency_pair_id: int,
    rate_type: str = "Market",
    start: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db),
    _u: models.User = Depends(get_current_user),
):
    pair = db.query(models.CurrencyPair).get(currency_pair_id)
    if not pair:
        raise HTTPException(404, "Unknown currency pair.")
    series = build_daily_series(db, pair.base_currency, pair.quote_currency, rate_type, start, end)
    rows = [
        {
            "Date": row["rate_date"],
            "Pair": pair.label,
            "Rate Type": rate_type,
            "Rate": float(row["rate"]),
            "Carried Forward": row["is_carried_forward"],
        }
        for row in series
    ]
    return xlsx_response(rows, f"fx_rates_{pair.label.replace('/', '')}_{rate_type}.xlsx")


@router.get("/import/template")
def fx_import_template(_u: models.User = Depends(get_current_user)):
    example_rows = [
        ["2026-08-24", "AED", "SDG", "Market", 981.20],
        ["2026-08-25", "AED", "SDG", "Market", 982.00],
        ["2026-08-25", "AED", "SDG", "CBOS", 970.00],
        ["2026-08-25", "AED", "SDG", "Pricing", 985.50],
        ["2026-08-25", "USD", "SDG", "Market", 3610.00],
    ]
    notes = [
        "Date must be YYYY-MM-DD (or any format Excel stores as a real date).",
        "Base Currency / Quote Currency: e.g. AED / SDG. Currency codes should already exist"
        " under Settings > Currencies, or be added there first -- unknown codes are skipped.",
        "Currency pairs are created automatically the first time they appear here if they"
        " don't exist yet.",
        "Rate Type must be exactly one of: Market, CBOS, Pricing. CBOS and Pricing are only"
        " valid for a pair where SDG is the base or quote currency -- rows using them for a"
        " non-SDG pair (e.g. USD/EUR) are skipped.",
        "Uploading a row for a Date + Pair + Rate Type that already has a rate on file UPDATES"
        " that rate (replaces the value) rather than creating a duplicate. This is how you"
        " correct a mistake or replace test data with final figures -- just re-upload with the"
        " corrected value.",
        "Rate is a number, e.g. 982.00 or 3,610.00 -- thousands-separator commas are fine, just"
        " don't include a currency symbol.",
    ]
    return xlsx_template_response(FX_TEMPLATE_COLUMNS, example_rows, notes, "fx_rates_template.xlsx")


@router.post("/import", response_model=schemas.ImportResult)
def fx_import(
    file: UploadFile,
    db: Session = Depends(get_db),
    _u: models.User = Depends(require_write),
):
    try:
        df = read_uploaded_xlsx(file.file.read())
    except Exception as e:
        raise HTTPException(400, f"Couldn't read this file as an Excel workbook: {e}")

    missing = [c for c in FX_TEMPLATE_COLUMNS if c not in df.columns]
    if missing:
        raise HTTPException(
            400,
            f"Missing column(s): {', '.join(missing)}. Download the template to see the"
            " expected columns.",
        )

    imported = 0
    updated = 0
    errors: list[schemas.ImportRowError] = []
    pair_cache: dict[tuple, models.CurrencyPair] = {}

    for idx, row in df.iterrows():
        row_number = idx + 2  # header is row 1, pandas is 0-indexed
        try:
            raw_date = row["Date"]
            try:
                rate_date = (
                    raw_date.date() if hasattr(raw_date, "date") else pd.to_datetime(raw_date).date()
                )
            except Exception:
                raise ValueError(
                    f"'{raw_date}' in the Date column isn't a valid calendar date -- this usually"
                    " happens when a date column is stored as text and Excel's fill-handle just"
                    " increments the trailing number (e.g. ...-08-31 becomes ...-08-32 instead of"
                    " rolling over to September). Format the column as a real Date and re-enter it."
                )
            base = str(row["Base Currency"]).strip().upper()
            quote = str(row["Quote Currency"]).strip().upper()
            rate_type = _normalize_rate_type(row["Rate Type"])
            rate = parse_decimal(row["Rate"])
        except (ValueError, InvalidOperation, TypeError) as e:
            errors.append(schemas.ImportRowError(row_number=row_number, reason=f"Couldn't parse row: {e}"))
            continue

        if rate_type is None:
            errors.append(
                schemas.ImportRowError(
                    row_number=row_number,
                    reason=f"Rate Type '{row['Rate Type']}' isn't recognized -- must be Market, CBOS, or Pricing.",
                )
            )
            continue

        if not db.query(models.Currency).get(base) or not db.query(models.Currency).get(quote):
            errors.append(
                schemas.ImportRowError(
                    row_number=row_number,
                    reason=f"Unknown currency code ({base} or {quote}) -- add it under Settings first.",
                )
            )
            continue

        # Each DB write below runs inside its own SAVEPOINT (nested
        # transaction). Without this, a DB-level failure on one row (e.g. a
        # constraint we didn't anticipate) leaves the whole session unusable
        # for every row after it, and the entire import fails with an opaque
        # 500 instead of a clear per-row skip reason.
        cache_key = (base, quote)
        pair = pair_cache.get(cache_key) or get_pair(db, base, quote)
        if not pair:
            try:
                with db.begin_nested():
                    pair = models.CurrencyPair(
                        base_currency=base,
                        quote_currency=quote,
                        supports_extended_rates="SDG" in (base, quote),
                        is_default=False,
                    )
                    db.add(pair)
                    db.flush()
            except Exception as e:
                errors.append(
                    schemas.ImportRowError(
                        row_number=row_number,
                        reason=f"Couldn't create currency pair {base}/{quote}: {e}",
                    )
                )
                continue
        pair_cache[cache_key] = pair

        allowed_types = SDG_RATE_TYPES if pair.supports_extended_rates else NON_SDG_RATE_TYPES
        if rate_type not in allowed_types:
            errors.append(
                schemas.ImportRowError(
                    row_number=row_number,
                    reason=f"Rate Type '{rate_type}' isn't valid for {pair.label} (allowed: {', '.join(allowed_types)}).",
                )
            )
            continue

        try:
            with db.begin_nested():
                existing = (
                    db.query(models.FxRate)
                    .filter(
                        models.FxRate.rate_date == rate_date,
                        models.FxRate.currency_pair_id == pair.id,
                        models.FxRate.rate_type == rate_type,
                    )
                    .first()
                )
                if existing:
                    existing.rate = rate
                    existing.is_manual_entry = True
                    row_was_update = True
                else:
                    db.add(
                        models.FxRate(
                            rate_date=rate_date,
                            currency_pair_id=pair.id,
                            rate_type=rate_type,
                            rate=rate,
                            is_manual_entry=True,
                        )
                    )
                    row_was_update = False
        except Exception as e:
            errors.append(
                schemas.ImportRowError(row_number=row_number, reason=f"Couldn't save row: {e}")
            )
            continue

        if row_was_update:
            updated += 1
        else:
            imported += 1

    db.commit()
    return schemas.ImportResult(imported=imported, updated=updated, skipped=errors)


# ---------------------------------------------------------------------------
# One-time historical seed: the "wide" layout the business's own currency
# table comes in (one row per calendar day, Market/CBOS/Pricing x
# USD/Euro/AED as separate columns) rather than the long Date/Pair/Rate
# Type/Rate layout the generic importer above expects. Header names are
# matched by text, not fixed column letters, so this tolerates the sheet's
# leading blank column and the blank spacer column between the AED-only and
# per-currency CBOS/Pricing sections.

_HISTORY_COLUMNS = {
    "date": "Date",
    "usd_market": "USD Rate",
    "euro_market": "Euro Rate",
    "aed_market": "AED Rate",
    "cbos_aed": "CBOS Rate (AED)",
    "pricing_aed": "Pricing Rate (AED)",
    "cbos_usd": "CBOS USD",
    "cbos_euro": "CBOS Euro",
    "pricing_usd": "Pricing USD",
    "pricing_euro": "Pricing Euro",
}


def _find_history_header_row(ws) -> tuple[dict[str, int], int] | None:
    for row in ws.iter_rows(min_row=1, max_row=20):
        by_text = {
            str(cell.value).strip().lower(): cell.column for cell in row if cell.value is not None
        }
        if "date" not in by_text:
            continue
        mapping: dict[str, int] = {}
        ok = True
        for key, header in _HISTORY_COLUMNS.items():
            col = by_text.get(header.strip().lower())
            if col is None:
                ok = False
                break
            mapping[key] = col
        if ok:
            return mapping, row[0].row
    return None


@router.get("/import-history/template")
def fx_import_history_template(_u: models.User = Depends(get_current_user)):
    columns = list(_HISTORY_COLUMNS.values())
    example_rows = [
        ["2026-01-01", 3610.08, 4332.10, 981.00, 750.00, 1034.58, 2754.75, 3305.70, 3800.00, 1241.49],
        ["2026-01-02", 3610.08, 4332.10, 981.00, 750.00, 1034.58, 2754.75, 3305.70, 3800.00, 1241.49],
    ]
    notes = [
        "One-time historical seed for the FX Rates page's Market/CBOS/Pricing tables -- NOT the"
        " same layout as the regular 'Import Rates from Excel' tool.",
        "One row per calendar day. All 10 columns must be present, with these exact header names"
        " (any column order is fine, and extra columns like 'Month' or 'Currency' are ignored).",
        "All 9 rate values are required on every row -- Market/CBOS/Pricing all require USD, Euro"
        " and AED together, same rule as manual entry on the FX Rates page.",
        "Re-uploading a Date that already has rates on file UPDATES those rates rather than"
        " creating duplicates -- safe to re-run after fixing a mistake.",
        "USD/SDG and EUR/SDG pairs are created automatically if they don't exist yet (AED/SDG"
        " already exists by default).",
    ]
    return xlsx_template_response(columns, example_rows, notes, "fx_history_import_template.xlsx")


@router.post("/import-history", response_model=schemas.ImportResult)
def fx_import_history(
    file: UploadFile,
    db: Session = Depends(get_db),
    _u: models.User = Depends(require_write),
):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Couldn't read this file as an Excel workbook: {e}")
    ws = wb[wb.sheetnames[0]]

    found = _find_history_header_row(ws)
    if not found:
        raise HTTPException(
            400,
            "Couldn't find the expected columns in this file. Expected: "
            + ", ".join(_HISTORY_COLUMNS.values())
            + ". Download the template to see the expected layout.",
        )
    col_map, header_row = found

    # --- Phase 1: parse the whole sheet in memory, no DB access at all. ---
    # A 200+ row file means up to 9 values/row to validate; doing that as
    # plain Python first (instead of interleaved with DB writes, as this
    # endpoint originally did) is what makes phase 2 below possible to run
    # as a handful of queries instead of thousands.
    errors: list[schemas.ImportRowError] = []
    parsed_rows: list[tuple[object, dict]] = []  # (rate_date, {key: Decimal})
    for r in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(row=r, column=col_map["date"]).value
        if raw_date is None:
            continue  # trailing/blank spacer row -- not an error, just skip

        try:
            rate_date = raw_date.date() if hasattr(raw_date, "date") else pd.to_datetime(raw_date).date()
        except Exception:
            errors.append(schemas.ImportRowError(row_number=r, reason=f"'{raw_date}' isn't a valid date."))
            continue

        try:
            values = {
                key: parse_decimal(ws.cell(row=r, column=col).value)
                for key, col in col_map.items()
                if key != "date"
            }
        except (InvalidOperation, TypeError, ValueError) as e:
            errors.append(
                schemas.ImportRowError(
                    row_number=r,
                    reason=f"{rate_date}: missing or invalid rate value ({e}) -- all 9 rate columns"
                    " are required for every date.",
                )
            )
            continue

        parsed_rows.append((rate_date, values))

    imported = 0
    updated = 0

    if parsed_rows:
        # --- Phase 2: one bulk write pass. ---
        # The original version issued a SELECT + a SAVEPOINT per one of the
        # 9 values on every row -- ~3,800+ individual round trips for a
        # 212-day file. That's cheap against a local Postgres with
        # near-zero latency (where this was first tested and looked
        # instant), but against a real hosted DB the per-round-trip
        # latency adds up to minutes -- long enough to hit a proxy/browser
        # timeout and report "failed" even though the writes kept going
        # and eventually committed, which is exactly what was observed in
        # production. This phase instead does a handful of queries total:
        # get-or-create the 3 pairs, ONE query to preload every existing
        # FxRate row in the affected range, then pure in-memory
        # add/update against the session before a single final commit.
        usd_pair = _get_or_create_sdg_pair(db, "USD")
        eur_pair = _get_or_create_sdg_pair(db, "EUR")
        aed_pair = _get_or_create_sdg_pair(db, "AED")
        db.flush()

        pair_ids = [usd_pair.id, eur_pair.id, aed_pair.id]
        dates = [d for d, _ in parsed_rows]
        existing_rows = (
            db.query(models.FxRate)
            .filter(
                models.FxRate.currency_pair_id.in_(pair_ids),
                models.FxRate.rate_type.in_(SDG_RATE_TYPES),
                models.FxRate.rate_date >= min(dates),
                models.FxRate.rate_date <= max(dates),
            )
            .all()
        )
        existing_by_key = {(row.currency_pair_id, row.rate_type, row.rate_date): row for row in existing_rows}

        try:
            for rate_date, values in parsed_rows:
                entries = [
                    (usd_pair, "Market", values["usd_market"]),
                    (eur_pair, "Market", values["euro_market"]),
                    (aed_pair, "Market", values["aed_market"]),
                    (aed_pair, "CBOS", values["cbos_aed"]),
                    (usd_pair, "CBOS", values["cbos_usd"]),
                    (eur_pair, "CBOS", values["cbos_euro"]),
                    (aed_pair, "Pricing", values["pricing_aed"]),
                    (usd_pair, "Pricing", values["pricing_usd"]),
                    (eur_pair, "Pricing", values["pricing_euro"]),
                ]
                for pair, rate_type, rate in entries:
                    key = (pair.id, rate_type, rate_date)
                    existing = existing_by_key.get(key)
                    if existing:
                        existing.rate = rate
                        existing.is_manual_entry = True
                        updated += 1
                    else:
                        new_row = models.FxRate(
                            rate_date=rate_date,
                            currency_pair_id=pair.id,
                            rate_type=rate_type,
                            rate=rate,
                            is_manual_entry=True,
                        )
                        db.add(new_row)
                        existing_by_key[key] = new_row
                        imported += 1
            db.commit()
        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Couldn't save this import -- nothing was written: {e}")

    return schemas.ImportResult(imported=imported, updated=updated, skipped=errors)
